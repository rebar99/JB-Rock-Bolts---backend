import csv
import json
import os
import uuid
from datetime import datetime
from io import BytesIO
from typing import Any, Dict, List
from fastapi import UploadFile

UPLOAD_DIR = "uploads"


def ensure_upload_dir(path: str) -> None:
    if not os.path.exists(path):
        os.makedirs(path, exist_ok=True)


async def read_upload_bytes(file: UploadFile) -> bytes:
    return await file.read()


def save_upload_bytes(content: bytes, filename: str, subfolder: str = "") -> str:
    ensure_upload_dir(UPLOAD_DIR)
    target_dir = os.path.join(UPLOAD_DIR, subfolder) if subfolder else UPLOAD_DIR
    ensure_upload_dir(target_dir)

    file_extension = os.path.splitext(filename)[1]
    unique_filename = f"{uuid.uuid4()}{file_extension}"
    file_path = os.path.join(target_dir, unique_filename)

    with open(file_path, "wb") as buffer:
        buffer.write(content)

    relative_url = f"/uploads/{subfolder}/{unique_filename}" if subfolder else f"/uploads/{unique_filename}"
    return relative_url


def _parse_xlsx(content: bytes) -> List[Dict[str, Any]]:
    import openpyxl
    wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(cell).strip() if cell is not None else f"col_{i}" for i, cell in enumerate(rows[0])]
    result = []
    for row in rows[1:]:
        if all(cell is None for cell in row):
            continue
        row_dict = {}
        for header, value in zip(headers, row):
            if header and header != "None":
                row_dict[header] = value if value is not None else ""
        if any(v not in (None, "") for v in row_dict.values()):
            result.append(row_dict)
    return result


def parse_import_file(content: bytes, filename: str) -> List[Dict[str, Any]]:
    extension = os.path.splitext(filename)[1].lower()

    if extension in (".xlsx", ".xls"):
        return _parse_xlsx(content)

    text = content.decode("utf-8-sig")

    if extension == ".json" or (not extension and text.strip().startswith("{")) or (not extension and text.strip().startswith("[")):
        parsed = json.loads(text)
        if isinstance(parsed, list):
            return parsed
        if isinstance(parsed, dict):
            return [parsed]
        raise ValueError("JSON import file must contain an object or an array of objects.")

    if extension in (".csv", ".txt") or not extension:
        rows: List[Dict[str, Any]] = []
        reader = csv.DictReader(text.splitlines())
        for row in reader:
            cleaned = {k: (v.strip() if isinstance(v, str) else v) for k, v in row.items()}
            rows.append(cleaned)
        return rows

    raise ValueError("Unsupported import file type. Use .xlsx, .csv, or .json.")


def parse_optional_datetime(value: Any) -> Any:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, str):
        for fmt in ("%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y"):
            try:
                return datetime.strptime(value.strip(), fmt)
            except ValueError:
                continue
    return None


def make_excel_response(wb, filename: str):
    """Return a Response with the full Excel file bytes.

    StreamingResponse + BytesIO breaks in uvicorn's async event loop when called
    from a synchronous route handler (uvicorn runs it in a thread pool, then
    tries to async-iterate the BytesIO — which is not async-iterable).
    Using Response(content=bytes) is simpler and always works.
    """
    from fastapi.responses import Response
    buffer = BytesIO()
    wb.save(buffer)
    return Response(
        content=buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def style_header_row(ws, col_count: int):
    """Apply bold blue header styling to row 1 of a worksheet."""
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 28
    for col_idx in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        ws.column_dimensions[get_column_letter(col_idx)].width = 20
