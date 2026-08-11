from fastapi import APIRouter, Depends, HTTPException, File, UploadFile, Query
from sqlalchemy.orm import Session, joinedload
from typing import Optional
from datetime import datetime
from app.database import get_db
from app.models.models import WorkOrder, WorkOrderSale
from app.schemas.work_order_reports import (
    WorkOrderReportRow, WorkOrderReportOut, WorkOrderSaleReportRow, WorkOrderSaleReportOut,
)
from app.routers.upload_helpers import make_excel_response, style_header_row, read_upload_bytes
from app.utils.helpers import compute_sale_taxable_and_gst, compute_sale_grand_total

router = APIRouter(prefix="/api/work-order-reports", tags=["Work Order Reports"])


def _filtered_work_orders(
    db: Session,
    from_date: Optional[datetime],
    to_date: Optional[datetime],
    client: Optional[str],
    project: Optional[str],
    wo_status: Optional[str],
):
    q = db.query(WorkOrder).options(joinedload(WorkOrder.line_items), joinedload(WorkOrder.work_order_sales).joinedload(WorkOrderSale.items))
    if from_date:
        q = q.filter(WorkOrder.created_at >= from_date)
    if to_date:
        from datetime import time
        to_date_end = datetime.combine(to_date.date(), time(23, 59, 59))
        q = q.filter(WorkOrder.created_at <= to_date_end)
    if client and client.lower() != "all":
        q = q.filter(WorkOrder.client_name.ilike(f"%{client}%"))
    if project and project.lower() != "all":
        q = q.filter(WorkOrder.project.ilike(f"%{project}%"))
    if wo_status and wo_status.lower() != "all":
        q = q.filter(WorkOrder.status == wo_status)
    return q.order_by(WorkOrder.created_at.desc()).all()


@router.get("", response_model=WorkOrderReportOut)
def get_work_order_report(
    from_date: Optional[datetime] = None,
    to_date: Optional[datetime] = None,
    client: Optional[str] = None,
    project: Optional[str] = None,
    status: Optional[str] = None,
    db: Session = Depends(get_db),
):
    orders = _filtered_work_orders(db, from_date, to_date, client, project, status)

    rows = []
    completed = pending = sold = 0
    for o in orders:
        total_q = o.total_qty
        completed_q = o.completed_qty
        pending_q = o.pending_quantity
        # Sales Quantity = actual quantity dispatched via Work Order Sales
        # (real WorkOrderSaleItem rows), not a heuristic — mirrors how the
        # Purchase Order Report's numbers always come from real Sale records.
        sales_q = sum(item.quantity for s in o.work_order_sales for item in s.items)

        if o.status == "Completed":
            completed += 1
        elif o.status == "Pending":
            pending += 1
        if o.work_order_sales:
            sold += 1

        rows.append(WorkOrderReportRow(
            id=o.id,
            wo_number=o.wo_number,
            client_name=o.client_name,
            project=o.project or "—",
            wo_date=o.wo_date.strftime("%d-%m-%Y") if o.wo_date else (o.created_at.strftime("%d-%m-%Y") if o.created_at else "—"),
            item=o.items_display,
            total_quantity=round(total_q, 2),
            completed_quantity=round(completed_q, 2),
            pending_quantity=round(pending_q, 2),
            sales_quantity=round(sales_q, 2),
            status=o.status or "Pending",
            uom=o.uom or "Nos",
            subtotal=round(o.subtotal, 2),
            gst_amount=round(o.gst_amount, 2),
            grand_total=round(o.grand_total, 2),
        ))

    return WorkOrderReportOut(
        rows=rows,
        total_work_orders=len(rows),
        completed_work_orders=completed,
        pending_work_orders=pending,
        sales_work_orders=sold,
    )


# ── Work Order Sales report (mirrors the Purchase Order Report's Sales tab) ──

@router.get("/sales", response_model=WorkOrderSaleReportOut)
def get_work_order_sales_report(
    from_date: Optional[datetime] = None,
    to_date: Optional[datetime] = None,
    client: Optional[str] = None,
    limit: int = 1000,
    db: Session = Depends(get_db),
):
    q = db.query(WorkOrderSale).options(joinedload(WorkOrderSale.items))
    if from_date:
        q = q.filter(WorkOrderSale.created_at >= from_date)
    if to_date:
        from datetime import time
        to_date_end = datetime.combine(to_date.date(), time(23, 59, 59))
        q = q.filter(WorkOrderSale.created_at <= to_date_end)
    if client and client.lower() != "all":
        q = q.filter(WorkOrderSale.client_name.ilike(f"%{client}%"))

    sales = q.order_by(WorkOrderSale.created_at.desc()).limit(limit).all()

    rows = []
    total_revenue = 0
    for s in sales:
        taxable_amount, gst_amount = compute_sale_taxable_and_gst(s.items)
        row_price = compute_sale_grand_total(taxable_amount, gst_amount, float(s.freight or 0))
        total_revenue += row_price

        rows.append(WorkOrderSaleReportRow(
            id=s.id,
            date=s.invoice_date.strftime("%d-%m-%Y") if s.invoice_date else (s.created_at.strftime("%d-%m-%Y") if s.created_at else ""),
            invoice_number=s.invoice_number,
            wo_number=s.wo_number,
            client_name=s.client_name,
            subtotal=round(taxable_amount, 2),
            gst_amount=round(gst_amount, 2),
            grand_total=round(row_price, 2),
            payment_status=s.payment_status.value if hasattr(s.payment_status, 'value') else str(s.payment_status),
        ))

    record_count = len(sales)
    avg_order_value = total_revenue / record_count if record_count else 0

    return WorkOrderSaleReportOut(
        rows=rows,
        total_revenue=round(total_revenue, 2),
        record_count=record_count,
        avg_order_value=round(avg_order_value, 2),
    )


# ── Combined multi-sheet export ───────────────────────────────────────────────

@router.get("/export-combined")
def export_combined_work_order_report(
    sheets: str = Query("completed,sales,pending", description="Comma-separated: completed, sales, pending"),
    from_date: Optional[datetime] = None,
    to_date: Optional[datetime] = None,
    client: Optional[str] = None,
    db: Session = Depends(get_db),
):
    """Export one or more Work Order report sections into a single Excel workbook (one sheet per section)."""
    from openpyxl import Workbook

    requested = [s.strip() for s in sheets.split(",") if s.strip()]
    if not requested:
        raise HTTPException(status_code=400, detail="No sheet types specified")

    wb = Workbook()
    wb.remove(wb.active)  # drop the default empty sheet

    if "completed" in requested:
        ws = wb.create_sheet("Completed WO Report")
        headers = ["Date", "Client Name", "Project", "WO Number", "Items",
                   "WO Quantity", "Completed Quantity", "UOM"]
        ws.append(headers)
        style_header_row(ws, len(headers))
        data = get_work_order_report(from_date=from_date, to_date=to_date, client=client, project=None, status=None, db=db)
        for r in data.rows:
            if r.pending_quantity > 0.0001:
                continue
            ws.append([r.wo_date, r.client_name, r.project, r.wo_number, r.item,
                       r.total_quantity, r.completed_quantity, r.uom])

    if "sales" in requested:
        ws = wb.create_sheet("Sales Report")
        headers = ["Date", "Invoice Number", "WO Number", "Client Name",
                   "Subtotal", "GST Amount", "Grand Total", "Payment Status"]
        ws.append(headers)
        style_header_row(ws, len(headers))
        data = get_work_order_sales_report(from_date=from_date, to_date=to_date, client=client, db=db)
        for r in data.rows:
            ws.append([r.date, r.invoice_number or "—", r.wo_number or "—", r.client_name,
                       r.subtotal, r.gst_amount, r.grand_total, r.payment_status])
        ws.append([])
        ws.append(["", "", "", "Total Revenue:", data.total_revenue])
        ws.append(["", "", "", "Record Count:", data.record_count])
        ws.append(["", "", "", "Avg Order Value:", data.avg_order_value])

    if "pending" in requested:
        ws = wb.create_sheet("Pending WOs Report")
        headers = ["WO Date", "WO Number", "Client Name", "Project", "Item",
                   "Total Qty", "Completed Qty", "Pending Qty", "UOM", "Status"]
        ws.append(headers)
        style_header_row(ws, len(headers))
        data = get_work_order_report(from_date=from_date, to_date=to_date, client=client, project=None, status=None, db=db)
        for r in data.rows:
            if r.pending_quantity <= 0.0001:
                continue
            ws.append([r.wo_date, r.wo_number, r.client_name, r.project, r.item,
                       r.total_quantity, r.completed_quantity, r.pending_quantity, r.uom, r.status])

    if not wb.sheetnames:
        raise HTTPException(status_code=400, detail="No valid sheet types in request")

    sheet_tag = "-".join(s[:4] for s in requested)
    return make_excel_response(wb, f"work-order-reports-combined-{sheet_tag}.xlsx")


@router.get("/export")
def export_work_order_report(
    from_date: Optional[datetime] = None,
    to_date: Optional[datetime] = None,
    client: Optional[str] = None,
    project: Optional[str] = None,
    status: Optional[str] = None,
    db: Session = Depends(get_db),
):
    from openpyxl import Workbook

    orders = _filtered_work_orders(db, from_date, to_date, client, project, status)

    wb = Workbook()
    ws = wb.active
    ws.title = "Work Order Report"

    headers = [
        "Work Order No.", "Client", "Project", "Work Order Date", "Item",
        "Total Quantity", "Completed Quantity", "Pending Quantity", "Sales Quantity", "Status",
    ]
    ws.append(headers)
    style_header_row(ws, len(headers))

    for o in orders:
        sales_q = sum(item.quantity for s in o.work_order_sales for item in s.items)
        ws.append([
            o.wo_number, o.client_name, o.project or "",
            o.wo_date.strftime("%Y-%m-%d") if o.wo_date else "",
            o.items_display,
            round(o.total_qty, 2), round(o.completed_qty, 2), round(o.pending_quantity, 2), round(sales_q, 2),
            o.status or "Pending",
        ])

    return make_excel_response(wb, "work-order-report-export.xlsx")


# ── Combined multi-sheet import (backup restore) ──────────────────────────────

@router.post("/import-combined")
async def import_combined_work_order_report(
    file: UploadFile = File(...),
    on_conflict: str = Query("skip", description="skip | update"),
    created_by: Optional[str] = None,
    db: Session = Depends(get_db),
):
    """Import a combined Work Order backup Excel file.

    Detects sheets by name and routes each to the appropriate importer:
      "Work Orders"       → Work Orders module
      "Work Order Sales"  → Work Order Sales module
    """
    from openpyxl import load_workbook
    from io import BytesIO
    from openpyxl import Workbook as _WB
    from app.routers.work_orders import import_work_orders
    from app.routers.work_order_sales import import_work_order_sales

    content = await read_upload_bytes(file)
    try:
        wb = load_workbook(BytesIO(content), data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Cannot read Excel file: {exc}")

    SHEET_ROUTES = {
        "Work Orders": "work-orders",
        "Work Order Sales": "work-order-sales",
        "Completed WO Report": "work-orders",
        "Pending WOs Report": "work-orders",
        "Sales Report": "work-order-sales",
    }

    class _FakeUpload:
        def __init__(self, data: bytes, name: str):
            self._data = data
            self.filename = name

        async def read(self) -> bytes:
            return self._data

    found = [s for s in wb.sheetnames if s in SHEET_ROUTES]
    if not found:
        raise HTTPException(
            status_code=400,
            detail=(f"No recognizable sheets found. Got: {wb.sheetnames}. "
                    f"Expected one or more of: {list(SHEET_ROUTES.keys())}"),
        )

    results: dict = {}
    for sheet_name in found:
        # Re-package just this sheet into its own mini workbook
        mini_wb = _WB()
        mini_ws = mini_wb.active
        mini_ws.title = sheet_name
        for row in wb[sheet_name].iter_rows(values_only=True):
            mini_ws.append(list(row))
        buf = BytesIO()
        mini_wb.save(buf)

        fake = _FakeUpload(buf.getvalue(), f"{sheet_name}.xlsx")
        if SHEET_ROUTES[sheet_name] == "work-order-sales":
            result = await import_work_order_sales(file=fake, on_conflict=on_conflict, created_by=created_by, db=db)
        else:
            result = await import_work_orders(file=fake, on_conflict=on_conflict, created_by=created_by, db=db)
        results[sheet_name] = result

    all_errors = [f"[{sn}] {e}" for sn, r in results.items() for e in (r.get("errors") or [])]
    return {
        "sheets": results,
        "created": sum(r.get("created", 0) for r in results.values()),
        "updated": sum(r.get("updated", 0) for r in results.values()),
        "skipped": sum(r.get("skipped", 0) for r in results.values()),
        "errors": all_errors,
    }
