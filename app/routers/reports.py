import re
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from sqlalchemy.orm import Session, joinedload
from typing import Optional
from datetime import datetime
from app.database import get_db
from app.models.models import Record, PurchaseOrder, Sale, POLineItem, SaleDispatch, WorkOrder
from app.schemas.reports import (
    ReportOut, ReportRow, FulfillmentReportOut, FulfillmentReportRow,
    PendingPOReportOut, PendingPORow, POFulfillmentSummaryOut, DispatchHistoryRow,
    OverviewReportOut, OverviewSummary, OverviewRow, OverviewProductSummary, OverviewSizeBreakdown,
)
from app.schemas.purchase_order import PurchaseOrderOut
from app.routers.upload_helpers import (
    read_upload_bytes, parse_import_file, make_excel_response, style_header_row,
)
from app.utils.helpers import (
    compute_sale_taxable_and_gst, compute_sale_grand_total, compute_line_taxable_and_gst,
    parse_item_type_and_size, normalize_client_name, normalize_project_name,
    dedupe_names_by_normalized_key,
)

router = APIRouter(prefix="/api/reports", tags=["Reports"])


# ── Overview Report (live product-wise pending dashboard) ────────────────────

@router.get("/overview", response_model=OverviewReportOut)
def get_overview_report(db: Session = Depends(get_db)):
    """Live snapshot of Ordered/Dispatched/Pending quantity across every
    Purchase Order and Work Order line item. Product Type and Dia/Pipe Size
    are not stored columns anywhere — they're derived on the fly from the
    free-typed item name via parse_item_type_and_size(), so a brand-new
    product name automatically gets its own group with no code change.

    No caching, no stored aggregate: every number here is summed straight
    from POLineItem.quantity/delivered_quantity and WOLineItem.quantity/
    completed_quantity on every request — the same live columns the PO/WO
    Report tabs already read — so this always reflects the current DB state.
    """
    pos = db.query(PurchaseOrder).options(joinedload(PurchaseOrder.line_items)).all()
    wos = db.query(WorkOrder).options(joinedload(WorkOrder.line_items)).all()

    rows: list[OverviewRow] = []

    for po in pos:
        for li in po.line_items:
            ordered = round(float(li.quantity or 0), 10)
            dispatched = round(float(li.delivered_quantity or 0), 10)
            pending = round(max(0.0, ordered - dispatched), 10)
            product_type, size, size_label = parse_item_type_and_size(li.item)
            rows.append(OverviewRow(
                source="PO",
                order_no=po.po_number,
                client_name=po.client_name,
                client_key=normalize_client_name(po.client_name) or (po.client_name or ""),
                project=po.project,
                product_type=product_type,
                item=li.item,
                uom=li.uom or "Nos",
                ordered_qty=ordered,
                dispatched_qty=dispatched,
                pending_qty=pending,
                size=size,
                size_label=size_label,
            ))

    for wo in wos:
        for li in wo.line_items:
            ordered = round(float(li.quantity or 0), 10)
            dispatched = round(float(li.completed_quantity or 0), 10)
            pending = round(max(0.0, ordered - dispatched), 10)
            product_type, size, size_label = parse_item_type_and_size(li.item)
            rows.append(OverviewRow(
                source="WO",
                order_no=wo.wo_number,
                client_name=wo.client_name,
                client_key=normalize_client_name(wo.client_name) or (wo.client_name or ""),
                project=wo.project,
                product_type=product_type,
                item=li.item,
                uom=li.uom or "Nos",
                ordered_qty=ordered,
                dispatched_qty=dispatched,
                pending_qty=pending,
                size=size,
                size_label=size_label,
            ))

    # Dynamic per-product-type summary — drives the Expand-1 size/dia
    # breakdown, grouped purely by whatever product_type strings were
    # derived above, so a new product automatically gets its own group
    # with no code change.
    product_totals: dict = {}
    size_totals: dict = {}  # (product_type, size) -> {..., "size_label": ...}
    for r in rows:
        pt = product_totals.setdefault(r.product_type, {"ordered": 0.0, "dispatched": 0.0, "pending": 0.0})
        pt["ordered"] += r.ordered_qty
        pt["dispatched"] += r.dispatched_qty
        pt["pending"] += r.pending_qty

    for po in pos:
        for li in po.line_items:
            product_type, size, size_label = parse_item_type_and_size(li.item)
            if size is None:
                continue
            ordered = round(float(li.quantity or 0), 10)
            dispatched = round(float(li.delivered_quantity or 0), 10)
            k = (product_type, size)
            st = size_totals.setdefault(k, {"ordered": 0.0, "dispatched": 0.0, "size_label": size_label})
            st["ordered"] += ordered
            st["dispatched"] += dispatched
    for wo in wos:
        for li in wo.line_items:
            product_type, size, size_label = parse_item_type_and_size(li.item)
            if size is None:
                continue
            ordered = round(float(li.quantity or 0), 10)
            dispatched = round(float(li.completed_quantity or 0), 10)
            k = (product_type, size)
            st = size_totals.setdefault(k, {"ordered": 0.0, "dispatched": 0.0, "size_label": size_label})
            st["ordered"] += ordered
            st["dispatched"] += dispatched

    def _size_sort_key(size: str):
        m = re.search(r"(\d+(?:\.\d+)?)", size)
        return float(m.group(1)) if m else 0.0

    products: list[OverviewProductSummary] = []
    for product_type, totals in sorted(product_totals.items(), key=lambda kv: kv[0]):
        breakdown = []
        for (pt, size), st in size_totals.items():
            if pt != product_type:
                continue
            ord_q = round(st["ordered"], 10)
            disp_q = round(st["dispatched"], 10)
            breakdown.append(OverviewSizeBreakdown(
                size=size,
                size_label=st["size_label"],
                ordered_qty=ord_q,
                dispatched_qty=disp_q,
                pending_qty=round(max(0.0, ord_q - disp_q), 10),
            ))
        breakdown.sort(key=lambda b: _size_sort_key(b.size))
        products.append(OverviewProductSummary(
            product_type=product_type,
            ordered_qty=round(totals["ordered"], 10),
            dispatched_qty=round(totals["dispatched"], 10),
            pending_qty=round(max(0.0, totals["pending"]), 10),
            size_breakdown=breakdown,
        ))

    total_clients = len({
        n for n in (normalize_client_name(po.client_name) for po in pos) if n
    } | {
        n for n in (normalize_client_name(wo.client_name) for wo in wos) if n
    })
    total_active_pos = sum(1 for po in pos if not po.short_closed and po.delivery_status != "Delivered")
    total_active_wos = sum(1 for wo in wos if wo.status not in ("Completed", "Closed", "Cancelled"))

    summary = OverviewSummary(
        total_ordered_qty=round(sum(r.ordered_qty for r in rows), 10),
        total_dispatched_qty=round(sum(r.dispatched_qty for r in rows), 10),
        total_pending_qty=round(sum(r.pending_qty for r in rows), 10),
        total_clients=total_clients,
        total_active_pos=total_active_pos,
        total_active_wos=total_active_wos,
        products=products,
    )

    rows.sort(key=lambda r: (r.client_name or "", r.product_type or "", r.item or ""))

    # Distinct project names per client, deduped the same way the PO/WO
    # project dropdown is (so "2952 Kochi Elevated Metro" and a jammed
    # "2952KochiElevatedMetro" typo collapse into one) — drives the
    # click-a-project step in the Overview client expand when a client has
    # 2+ projects on record. Split by PO vs WO (not one combined dict) so
    # the Work Order Overview never lists a project that only ever showed
    # up on that client's Purchase Orders, and vice versa.
    def _client_projects_for(source: str) -> dict:
        raw: dict = {}
        for r in rows:
            if r.source == source and r.project and r.project.strip():
                raw.setdefault(r.client_key, []).append(r.project.strip())
        return {
            client_key: dedupe_names_by_normalized_key(names, normalize_project_name)
            for client_key, names in raw.items()
        }

    client_projects = {
        "PO": _client_projects_for("PO"),
        "WO": _client_projects_for("WO"),
    }

    return OverviewReportOut(
        generated_at=datetime.utcnow(),
        summary=summary,
        rows=rows,
        client_projects=client_projects,
    )


# ── Sales Report ──────────────────────────────────────────────────────────────

@router.get("", response_model=ReportOut)
def get_report(
    from_date: Optional[datetime] = None,
    to_date: Optional[datetime] = None,
    product: Optional[str] = None,
    client: Optional[str] = None,
    limit: int = 1000,
    db: Session = Depends(get_db),
):
    q = db.query(Sale)
    if from_date:
        q = q.filter(Sale.created_at >= from_date)
    if to_date:
        from datetime import time
        to_date_end = datetime.combine(to_date.date(), time(23, 59, 59))
        q = q.filter(Sale.created_at <= to_date_end)
    if product:
        from app.models.models import SaleItem
        # Filter via a correlated EXISTS (Sale.items.any(...)) instead of an
        # explicit JOIN. A JOIN here — combined with the joinedload(Sale.items)
        # below — fans out one row per matching SaleItem times one row per
        # eager-loaded item, so a multi-item invoice gets counted (and its
        # GST/subtotal/price summed) more than once. any() adds a WHERE EXISTS
        # clause instead, which can't multiply the Sale row at all.
        q = q.filter(Sale.items.any(SaleItem.item.ilike(f"%{product}%")))
    if client and client.lower() != "all":
        q = q.filter(Sale.client_name.ilike(f"%{client}%"))

    sales = q.options(
        joinedload(Sale.items),
        joinedload(Sale.purchase_order),
    ).order_by(Sale.created_at.desc()).limit(limit).all()

    # Belt-and-suspenders: guarantee every Sale contributes to the report
    # exactly once, regardless of how the query above is written.
    seen_ids = set()
    deduped_sales = []
    for s in sales:
        if s.id not in seen_ids:
            seen_ids.add(s.id)
            deduped_sales.append(s)
    sales = deduped_sales

    # total_revenue is accumulated below from the exact same row_price values
    # that populate `rows` — no separate SQL SUM query. That guarantees the
    # "Filtered Revenue" summary card (which reads this field) can never
    # disagree with the table footer (which sums rows.price on the frontend):
    # they are, byte-for-byte, sums of the same numbers.
    rows = []
    total_revenue = 0
    for s in sales:
        po = s.purchase_order
        dispatched_qty = round(sum(item.quantity for item in s.items), 10)
        total_qty = round(float(po.total_qty if po else 0), 10)
        pending_qty = round(float(po.pending_quantity if po else 0), 10)

        # Fresh, independent GST calculation for this Sales record only:
        # Taxable Amount and GST Amount are derived straight from this sale's
        # own line items (quantity x unit_price, and x gst_rate/100) — never
        # from the stored Sale.subtotal/gst_amount aggregate columns. Grand
        # Total is then derived from these same two numbers (+ freight), the
        # same formula used everywhere else Grand Total is shown, so the
        # Sales Report can never disagree with the Sales page over GST.
        taxable_amount, gst_amount = compute_sale_taxable_and_gst(s.items)
        row_price = compute_sale_grand_total(taxable_amount, gst_amount, float(s.freight or 0))
        total_revenue += row_price

        rows.append(ReportRow(
            id=s.id,
            date=s.created_at.strftime("%d-%m-%Y") if s.created_at else "",
            client_name=s.client_name,
            product=s.items_display,
            location=s.project or "—",
            po_number=s.po_number,
            invoice_number=s.invoice_number,
            e_way_bill_no=s.e_way_bill_no,
            price=row_price,
            subtotal=taxable_amount,
            gst_amount=gst_amount,
            payment_status=s.payment_status.value if hasattr(s.payment_status, 'value') else str(s.payment_status),
            delivery_status="Dispatched",
            payment_note=s.payment_note,
            dispatched_qty=dispatched_qty,
            total_qty=total_qty,
            pending_qty=pending_qty,
            uom=po.uom if po else "Nos",
        ))

    record_count = len(sales)
    avg_order_value = total_revenue / record_count if record_count else 0

    return ReportOut(
        rows=rows,
        total_revenue=round(total_revenue, 2),
        record_count=record_count,
        avg_order_value=round(avg_order_value, 2),
    )


# ── Fulfillment Report ────────────────────────────────────────────────────────

@router.get("/fulfillment", response_model=FulfillmentReportOut)
def get_fulfillment_report(
    from_date: Optional[datetime] = None,
    to_date: Optional[datetime] = None,
    client: Optional[str] = None,
    db: Session = Depends(get_db),
):
    q = db.query(PurchaseOrder).options(joinedload(PurchaseOrder.line_items))
    if from_date:
        q = q.filter(PurchaseOrder.created_at >= from_date)
    if to_date:
        from datetime import time
        to_date_end = datetime.combine(to_date.date(), time(23, 59, 59))
        q = q.filter(PurchaseOrder.created_at <= to_date_end)
    if client and client.lower() != "all":
        q = q.filter(PurchaseOrder.client_name.ilike(f"%{client}%"))

    orders = q.order_by(PurchaseOrder.created_at.desc()).all()

    # Required/Delivered/Pending here are read from PurchaseOrderOut — the
    # exact same schema GET /api/purchase-orders uses to populate the
    # Purchase Orders table. This isn't "the same formula re-implemented":
    # it's the same computed_field code path, so there is no second place
    # these numbers could be calculated differently or drift out of sync.
    # Each PO is validated into PurchaseOrderOut exactly once (one row per
    # PurchaseOrder here, no JOIN to Sales/SaleItem that could duplicate it),
    # and PurchaseOrderOut.pending_quantity is itself total_quantity minus
    # delivered_quantity, so Required = Delivered + Pending holds for every
    # row — and therefore for the footer sums too.
    rows = []
    for o in orders:
        po_out = PurchaseOrderOut.model_validate(o)
        rows.append(FulfillmentReportRow(
            id=o.id,
            po_number=o.po_number,
            date=o.created_at.strftime("%d-%m-%Y") if o.created_at else "—",
            client_name=o.client_name,
            project=o.project or "—",
            item=o.items_display,
            total_required=po_out.total_quantity,
            delivered=po_out.delivered_quantity,
            pending=po_out.pending_quantity,
            uom=po_out.uom,
            remark=o.remark,
        ))

    return FulfillmentReportOut(rows=rows)


# ── Pending POs Report ────────────────────────────────────────────────────────

@router.get("/pending-pos", response_model=PendingPOReportOut)
def get_pending_pos_report(db: Session = Depends(get_db)):
    from app.models.models import DeliveryStatus

    all_pos = (
        db.query(PurchaseOrder)
        .options(joinedload(PurchaseOrder.sales).joinedload(Sale.items))
        .order_by(PurchaseOrder.created_at.desc())
        .all()
    )
    pending_orders = [o for o in all_pos if o.delivery_status != DeliveryStatus.DELIVERED]

    rows = []
    total_subtotal = total_gst = total_value = total_pending_value = total_delivered_payment = 0

    for o in pending_orders:
        t_sub = float(o.subtotal or 0)
        t_gst = float(o.gst_amount or 0)
        t_total = float(o.grand_total or 0)

        total_subtotal += t_sub
        total_gst += t_gst
        total_value += t_total

        t_qty = float(o.total_qty or 0)
        d_qty = float(o.delivered_qty or 0)

        # Delivered payment = sum of actual invoiced amounts against this PO,
        # each Sale calculated fresh via the same shared helpers as the Sales
        # Report/Dashboard — never the stored Sale.subtotal/gst_amount/
        # grand_total columns, which are only a snapshot from sale creation.
        delivered_sub = delivered_gst_amt = delivered_payment = 0.0
        for s in o.sales:
            s_taxable, s_gst = compute_sale_taxable_and_gst(s.items)
            delivered_sub += s_taxable
            delivered_gst_amt += s_gst
            delivered_payment += compute_sale_grand_total(s_taxable, s_gst, float(s.freight or 0))

        p_sub = max(0, t_sub - delivered_sub)
        p_gst = max(0, t_gst - delivered_gst_amt)
        p_total = max(0, t_total - delivered_payment)

        pending_qty = max(0, t_qty - d_qty)
        # Fallback: if pending qty exists but pending payment = 0 (PO price missing/wrong),
        # estimate from the per-unit rate of actual invoices
        if p_total == 0 and pending_qty > 0 and d_qty > 0 and delivered_payment > 0:
            invoice_rate = delivered_payment / d_qty
            p_total = round(invoice_rate * pending_qty, 2)

        total_pending_value += p_total
        total_delivered_payment += delivered_payment

        rows.append(PendingPORow(
            id=o.id,
            po_number=o.po_number,
            client_name=o.client_name,
            project=o.project or "—",
            item=o.items_display,
            subtotal=round(t_sub, 2),
            gst_amount=round(t_gst, 2),
            total_value=round(t_total, 2),
            pending_subtotal=round(p_sub, 2),
            pending_gst=round(p_gst, 2),
            pending_total=round(p_total, 2),
            delivered_payment=round(delivered_payment, 2),
            status=o.delivery_status,
            date=o.created_at.strftime("%d-%m-%Y") if o.created_at else "—",
            uom=o.uom or "Nos",
            total_qty=round(t_qty, 2),
            delivered_qty=round(d_qty, 2),
            pending_qty=round(pending_qty, 2),
            remark=o.remark,
        ))

    return PendingPOReportOut(
        rows=rows,
        total_subtotal=round(total_subtotal, 2),
        total_gst=round(total_gst, 2),
        total_value=round(total_value, 2),
        total_pending_value=round(total_pending_value, 2),
        total_delivered_payment=round(total_delivered_payment, 2),
        count=len(rows)
    )


# ── PO Fulfillment Summary (Reports → PO Fulfillment drill-down) ──────────────

@router.get("/po-fulfillment-summary/{po_id}", response_model=POFulfillmentSummaryOut)
def get_po_fulfillment_summary(po_id: int, db: Session = Depends(get_db)):
    po = (
        db.query(PurchaseOrder)
        .options(
            joinedload(PurchaseOrder.sales).joinedload(Sale.items),
            joinedload(PurchaseOrder.sales).joinedload(Sale.dispatches).joinedload(SaleDispatch.items),
        )
        .filter(PurchaseOrder.id == po_id)
        .first()
    )
    if not po:
        raise HTTPException(status_code=404, detail="Purchase order not found.")

    po_out = PurchaseOrderOut.model_validate(po)

    # Dispatch history: one row per ITEM per actual dispatch EVENT — a
    # dispatch covering two items (e.g. one measured in Meter, one in Nos)
    # shows as two rows here, each with its own item/qty/uom, instead of one
    # row with a combined quantity that silently mixes units together. An
    # invoice dispatched twice (same day or different days) via "Dispatch
    # More" likewise shows each event's items separately, tagged with that
    # event's own date. `total_dispatches`/`last_dispatch_date` count
    # dispatch EVENTS, not item-rows. Sales/dispatches recorded before
    # per-item tracking existed fall back to an aggregate row (or, for the
    # untracked remainder, a per-item-name estimate) — amounts for these
    # fallback rows are computed fresh from the sale's own line items (never
    # the stored Sale.grand_total snapshot column), the same shared helpers
    # used by the Sales Report and Pending POs Report, so this panel can
    # never disagree with those.
    dated_rows = []  # (event_timestamp, DispatchHistoryRow)
    payment_received = 0.0
    event_count = 0
    event_dates = []

    for s in po.sales:
        s_taxable, s_gst = compute_sale_taxable_and_gst(s.items)
        invoice_amount = compute_sale_grand_total(s_taxable, s_gst, float(s.freight or 0))

        payment_status = s.payment_status.value if hasattr(s.payment_status, "value") else str(s.payment_status)
        if payment_status == "Paid":
            payment_received += invoice_amount

        # The sale's CURRENT invoiced items are the source of truth for how
        # much of each item name this sale actually represents. Dispatch
        # events are an append-only log (each "Dispatch More" adds rows) that
        # never shrinks on its own when an item is later edited/reduced on
        # the invoice, so per-item dispatch quantity below is capped against
        # this budget — keyed case-insensitively since item names are
        # free-typed and casing isn't always consistent between dispatch
        # events (e.g. "Couplers" vs "couplers").
        def _key(name):
            return (name or "").strip().lower()

        item_qty_by_name = {}
        item_uom_by_name = {}
        item_value_by_name = {}
        item_display_name = {}
        for it in s.items:
            k = _key(it.item)
            item_qty_by_name[k] = item_qty_by_name.get(k, 0) + float(it.quantity)
            item_uom_by_name[k] = it.uom
            item_display_name.setdefault(k, it.item)
            it_taxable, it_gst = compute_line_taxable_and_gst(it.quantity, float(it.unit_price), float(it.gst_rate))
            item_value_by_name[k] = item_value_by_name.get(k, 0) + it_taxable + it_gst

        # How much of each item name is already accounted for by a recorded
        # dispatch event that has per-item detail, plus how much total
        # quantity is tracked by dispatches that *don't* (legacy aggregate
        # events, whose item composition is unknown) — both feed the
        # remainder pass below so it never double-counts quantity that an
        # aggregate-only dispatch already covered.
        tracked_qty_by_item = {}
        aggregate_only_tracked_qty = 0.0

        for d in s.dispatches:
            event_date = d.dispatched_at or datetime.min
            event_invoice = s.invoice_number  # Always use sale's invoice_number for consistency
            event_had_visible_row = False

            if d.items:
                # Merge same item+UOM lines *within this one dispatch event*
                # first (e.g. two SaleDispatchItem rows for "Couplers"/Nos
                # recorded in the same dispatch) so this event contributes a
                # single quantity per item — never a second, separate event
                # date for what was actually one dispatch.
                merged_items = {}
                merged_order = []
                for di in d.items:
                    mkey = (di.item, di.uom)
                    if mkey not in merged_items:
                        merged_items[mkey] = {"item": di.item, "uom": di.uom, "quantity": 0.0, "amount": 0.0}
                        merged_order.append(mkey)
                    merged_items[mkey]["quantity"] += float(di.quantity)
                    merged_items[mkey]["amount"] += float(di.amount)

                for mkey in merged_order:
                    mi = merged_items[mkey]
                    k = _key(mi["item"])
                    already = tracked_qty_by_item.get(k, 0)
                    remaining_budget = max(0.0, item_qty_by_name.get(k, 0) - already)
                    if remaining_budget <= 0:
                        # This item's quantity was fully removed/edited out of
                        # the invoice since this dispatch was recorded — drop
                        # the now-phantom row instead of showing quantity that
                        # no longer exists on the invoice.
                        continue
                    shown_qty = min(mi["quantity"], remaining_budget)
                    ratio = shown_qty / mi["quantity"] if mi["quantity"] else 0
                    tracked_qty_by_item[k] = already + shown_qty
                    event_had_visible_row = True
                    dated_rows.append((event_date, DispatchHistoryRow(
                        date=event_date.strftime("%d-%m-%Y") if d.dispatched_at else "—",
                        invoice_number=event_invoice,
                        item=mi["item"],
                        dispatch_qty=round(shown_qty, 10),
                        uom=mi["uom"],
                        amount=round(mi["amount"] * ratio, 2),
                        sale_id=s.id,
                        dispatch_id=d.id,
                    )))
            else:
                # Dispatch event recorded before per-item tracking existed —
                # only an aggregate qty/uom is available for it.
                aggregate_only_tracked_qty += float(d.quantity)
                event_had_visible_row = True
                dated_rows.append((event_date, DispatchHistoryRow(
                    date=event_date.strftime("%d-%m-%Y") if d.dispatched_at else "—",
                    invoice_number=event_invoice,
                    item=None,
                    dispatch_qty=round(float(d.quantity), 10),
                    uom=d.uom,
                    amount=round(float(d.amount), 2),
                    sale_id=s.id,
                    dispatch_id=d.id,
                )))

            # Only count this as one of "Total Dispatches" if it actually
            # left a visible row — a dispatch whose entire item quantity has
            # since been edited/removed out of the invoice (all rows dropped
            # above) shouldn't still be counted as a live dispatch event.
            if event_had_visible_row:
                event_count += 1
                event_dates.append(event_date)

        # Untracked remainder — per item name, so mixed-uom items are never
        # summed together. Covers legacy invoices with no SaleDispatch rows
        # at all, and legacy dispatch events recorded before per-item
        # tracking; without this the untracked quantity would silently
        # vanish from the history instead of showing as its own entry.
        remaining_after_known = {
            k: max(0.0, qty - tracked_qty_by_item.get(k, 0))
            for k, qty in item_qty_by_name.items()
        }
        pool = sum(remaining_after_known.values())
        scale = max(0.0, pool - aggregate_only_tracked_qty) / pool if pool > 0 else 1.0

        had_remainder = False
        for k, remaining_qty in remaining_after_known.items():
            remainder_qty = round(remaining_qty * scale, 10)
            if remainder_qty <= 0:
                continue
            had_remainder = True
            total_qty = item_qty_by_name[k]
            ratio = remainder_qty / total_qty if total_qty else 0
            dated_rows.append((s.created_at or datetime.min, DispatchHistoryRow(
                date=s.created_at.strftime("%d-%m-%Y") if s.created_at else "—",
                invoice_number=s.invoice_number,
                item=item_display_name[k],
                dispatch_qty=remainder_qty,
                uom=item_uom_by_name[k],
                amount=round(item_value_by_name[k] * ratio, 2),
            )))

        if had_remainder or not s.dispatches:
            event_count += 1
            event_dates.append(s.created_at or datetime.min)

    dated_rows.sort(key=lambda pair: pair[0])
    dispatch_rows = [row for _, row in dated_rows]
    last_dispatch_date = max(event_dates) if event_dates else None

    pending_payment = max(0.0, po_out.grand_total - payment_received)

    if po.short_closed:
        delivery_status = "Short Closed"
    elif po_out.delivered_quantity <= 0:
        delivery_status = "Pending"
    elif po_out.delivered_quantity >= po_out.total_quantity:
        delivery_status = "Completed"
    else:
        delivery_status = "Partial"

    return POFulfillmentSummaryOut(
        po_id=po.id,
        po_number=po.po_number,
        client_name=po.client_name,
        project=po.project,
        po_date=po.po_date.strftime("%d-%m-%Y") if po.po_date else None,
        po_quantity=round(po_out.total_quantity, 10),
        delivered_quantity=round(po_out.delivered_quantity, 10),
        pending_quantity=po_out.pending_quantity,
        uom=po_out.uom,
        subtotal=round(po_out.subtotal, 2),
        gst_amount=round(po_out.gst_amount, 2),
        grand_total=round(po_out.grand_total, 2),
        dispatch_history=dispatch_rows,
        total_dispatches=event_count,
        last_dispatch_date=last_dispatch_date.strftime("%d-%m-%Y") if last_dispatch_date else None,
        payment_received=round(payment_received, 2),
        pending_payment=round(pending_payment, 2),
        delivery_status=delivery_status,
    )


# ── Combined multi-sheet export ───────────────────────────────────────────────

@router.get("/export-combined")
def export_combined_report(
    sheets: str = Query("fulfillment,sales,pending-pos", description="Comma-separated: fulfillment, sales, pending-pos"),
    from_date: Optional[datetime] = None,
    to_date: Optional[datetime] = None,
    client: Optional[str] = None,
    product: Optional[str] = None,
    db: Session = Depends(get_db),
):
    """Export one or more report sections into a single Excel workbook (one sheet per section)."""
    from openpyxl import Workbook

    requested = [s.strip() for s in sheets.split(",") if s.strip()]
    if not requested:
        raise HTTPException(status_code=400, detail="No sheet types specified")

    wb = Workbook()
    wb.remove(wb.active)  # drop the default empty sheet

    if "completed" in requested:
        ws = wb.create_sheet("Completed PO")
        headers = ["Date", "Client Name", "Project", "PO Number", "Items",
                   "PO Quantity", "Delivered", "UOM"]
        ws.append(headers)
        style_header_row(ws, len(headers))
        data = get_fulfillment_report(from_date=from_date, to_date=to_date, client=client, db=db)
        for r in data.rows:
            if r.pending > 0.0001:
                continue
            ws.append([r.date, r.client_name, r.project, r.po_number, r.item,
                       r.total_required, r.delivered, r.uom])

    if "sales" in requested:
        ws = wb.create_sheet("Sales Report")
        headers = ["Date", "Client Name", "Project / Location", "Items",
                   "Invoice No", "PO No", "E-Way Bill No",
                   "Total Qty", "Dispatched Qty", "Pending Qty", "UOM",
                   "Grand Total (₹)", "Payment Status"]
        ws.append(headers)
        style_header_row(ws, len(headers))
        data = get_report(from_date=from_date, to_date=to_date, product=product, client=client, db=db)
        for r in data.rows:
            ws.append([r.date, r.client_name, r.location, r.product,
                       r.invoice_number or "—", r.po_number or "—", r.e_way_bill_no or "—",
                       r.total_qty, r.dispatched_qty, r.pending_qty, r.uom,
                       r.price, r.payment_status])
        ws.append([])
        ws.append(["", "", "", "", "", "", "Total Revenue:", data.total_revenue])
        ws.append(["", "", "", "", "", "", "Record Count:", data.record_count])
        ws.append(["", "", "", "", "", "", "Avg Order Value:", data.avg_order_value])

    if "pending-pos" in requested:
        ws = wb.create_sheet("Pending POs")
        headers = ["Date", "PO Number", "Client Name", "Project", "Items",
                   "Value (Excl. GST) ₹", "GST Amount ₹", "Total Value ₹",
                   "Delivered Payment ₹", "Pending Total ₹", "Status"]
        ws.append(headers)
        style_header_row(ws, len(headers))
        data = get_pending_pos_report(db=db)
        for r in data.rows:
            ws.append([r.date, r.po_number, r.client_name, r.project, r.item,
                       r.subtotal, r.gst_amount, r.total_value,
                       r.delivered_payment, r.pending_total,
                       str(r.status.value if hasattr(r.status, "value") else r.status)])
        ws.append([])
        ws.append(["", "", "", "", "Totals:",
                   data.total_subtotal, data.total_gst, data.total_value,
                   data.total_delivered_payment, data.total_pending_value])

    if not wb.sheetnames:
        raise HTTPException(status_code=400, detail="No valid sheet types in request")

    sheet_tag = "-".join(s[:3] for s in requested)
    return make_excel_response(wb, f"reports-combined-{sheet_tag}.xlsx")


# ── Combined multi-sheet import ───────────────────────────────────────────────

@router.post("/import-combined")
async def import_combined_report_data(
    file: UploadFile = File(...),
    on_conflict: str = Query("skip", description="skip | update"),
    created_by: Optional[str] = None,
    db: Session = Depends(get_db),
):
    """Import a combined report Excel file.

    Detects sheets by name and routes each to the appropriate importer:
      "PO Fulfillment Report" (or legacy "Fulfillment Report") → Purchase Orders
      "Pending POs"                                            → Purchase Orders
      "Sales Report"                                            → Sales Invoices
    """
    from openpyxl import load_workbook
    from io import BytesIO
    from openpyxl import Workbook as _WB
    from app.routers.purchase_orders import import_purchase_orders
    from app.routers.sales import import_sales

    content = await read_upload_bytes(file)
    try:
        wb = load_workbook(BytesIO(content), data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Cannot read Excel file: {exc}")

    SHEET_ROUTES = {
        "Fulfillment Report": "purchase-orders",
        "PO Fulfillment Report": "purchase-orders",
        "Completed PO": "purchase-orders",
        "Pending POs": "purchase-orders",
        "Sales Report": "sales",
    }

    class _FakeUpload:
        def __init__(self, data: bytes, name: str):
            self._data = data
            self.filename = name

        async def read(self) -> bytes:
            return self._data

    found = [s for s in wb.sheetnames if s in SHEET_ROUTES]
    if not found:
        raise HTTPException(
            status_code=400,
            detail=(f"No recognizable sheets found. Got: {wb.sheetnames}. "
                    f"Expected one or more of: {list(SHEET_ROUTES.keys())}"),
        )

    results: dict = {}
    for sheet_name in found:
        # Re-package just this sheet into its own mini workbook
        mini_wb = _WB()
        mini_ws = mini_wb.active
        mini_ws.title = sheet_name
        for row in wb[sheet_name].iter_rows(values_only=True):
            mini_ws.append(list(row))
        buf = BytesIO()
        mini_wb.save(buf)

        fake = _FakeUpload(buf.getvalue(), f"{sheet_name}.xlsx")
        if SHEET_ROUTES[sheet_name] == "sales":
            result = await import_sales(file=fake, on_conflict=on_conflict, created_by=created_by, db=db)
        else:
            result = await import_purchase_orders(file=fake, on_conflict=on_conflict, created_by=created_by, db=db)
        results[sheet_name] = result

    all_errors = [f"[{sn}] {e}" for sn, r in results.items() for e in (r.get("errors") or [])]
    return {
        "sheets": results,
        "created": sum(r.get("created", 0) for r in results.values()),
        "updated": sum(r.get("updated", 0) for r in results.values()),
        "skipped": sum(r.get("skipped", 0) for r in results.values()),
        "errors": all_errors,
    }


# ── Single-sheet Excel export ──────────────────────────────────────────────────

@router.get("/export")
def export_report(
    report_type: str = Query("fulfillment", description="fulfillment | sales | pending-pos"),
    from_date: Optional[datetime] = None,
    to_date: Optional[datetime] = None,
    client: Optional[str] = None,
    product: Optional[str] = None,
    db: Session = Depends(get_db),
):
    """Export a report section to Excel (.xlsx)."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active

    if report_type == "sales":
        ws.title = "Sales Report"
        headers = [
            "Date", "Client Name", "Project / Location", "Items",
            "Invoice No", "PO No", "E-Way Bill No",
            "Total Qty", "Dispatched Qty", "Pending Qty", "UOM",
            "Grand Total (₹)", "Payment Status",
        ]
        ws.append(headers)
        style_header_row(ws, len(headers))

        data = get_report(from_date=from_date, to_date=to_date, product=product, client=client, db=db)
        for r in data.rows:
            ws.append([
                r.date, r.client_name, r.location, r.product,
                r.invoice_number or "—", r.po_number or "—", r.e_way_bill_no or "—",
                r.total_qty, r.dispatched_qty, r.pending_qty, r.uom,
                r.price, r.payment_status,
            ])

        ws.append([])
        ws.append(["", "", "", "", "", "", "Total Revenue:", data.total_revenue])
        ws.append(["", "", "", "", "", "", "Record Count:", data.record_count])
        ws.append(["", "", "", "", "", "", "Avg Order Value:", data.avg_order_value])

        filename = "sales-report.xlsx"

    elif report_type == "pending-pos":
        ws.title = "Pending POs"
        headers = [
            "Date", "PO Number", "Client Name", "Project", "Items",
            "Value (Excl. GST) ₹", "GST Amount ₹", "Total Value ₹",
            "Delivered Payment ₹", "Pending Total ₹", "Status",
        ]
        ws.append(headers)
        style_header_row(ws, len(headers))

        data = get_pending_pos_report(db=db)
        for r in data.rows:
            ws.append([
                r.date, r.po_number, r.client_name, r.project, r.item,
                r.subtotal, r.gst_amount, r.total_value,
                r.delivered_payment, r.pending_total,
                str(r.status.value if hasattr(r.status, "value") else r.status),
            ])

        ws.append([])
        ws.append(["", "", "", "", "Totals:", data.total_subtotal, data.total_gst, data.total_value,
                   data.total_delivered_payment, data.total_pending_value])

        filename = "pending-pos-report.xlsx"

    else:  # fulfillment (default)
        ws.title = "PO Fulfillment Report"
        headers = [
            "Date", "Client Name", "Project", "PO Number", "Items",
            "PO Quantity", "Delivered", "Pending", "UOM",
        ]
        ws.append(headers)
        style_header_row(ws, len(headers))

        data = get_fulfillment_report(from_date=from_date, to_date=to_date, client=client, db=db)
        for r in data.rows:
            ws.append([
                r.date, r.client_name, r.project, r.po_number, r.item,
                r.total_required, r.delivered, r.pending, r.uom,
            ])

        filename = "fulfillment-report.xlsx"

    return make_excel_response(wb, filename)


# ── Excel import (delegates to PO / Sales import logic) ──────────────────────

@router.post("/import")
async def import_report_data(
    file: UploadFile = File(...),
    report_type: str = Query("fulfillment", description="fulfillment | sales | pending-pos"),
    on_conflict: str = Query("skip", description="skip | update"),
    created_by: Optional[str] = None,
    db: Session = Depends(get_db),
):
    """Import report data from an Excel or CSV file.

    • fulfillment / pending-pos → creates / updates Purchase Orders
    • sales                     → creates / updates Sales Invoices
    """
    from app.routers.purchase_orders import import_purchase_orders
    from app.routers.sales import import_sales

    # Re-wrap the already-consumed file bytes into a fake UploadFile-compatible object
    content = await read_upload_bytes(file)

    class _FakeUpload:
        def __init__(self, data: bytes, name: str):
            self._data = data
            self.filename = name

        async def read(self) -> bytes:
            return self._data

    fake = _FakeUpload(content, file.filename)

    if report_type == "sales":
        return await import_sales(file=fake, on_conflict=on_conflict, created_by=created_by, db=db)
    else:
        return await import_purchase_orders(file=fake, on_conflict=on_conflict, created_by=created_by, db=db)
